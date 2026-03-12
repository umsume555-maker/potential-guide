import sqlite3
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelExporter:
    def __init__(self, db_path: str):
        self.db_path = db_path

    def export_ocr_results(self, run_id: str, base_url: str = "http://localhost:8000") -> BytesIO:
        """OCR結果をExcelファイルとしてエクスポート"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "請求書OCR結果"
        
        # ヘッダー定義
        headers = [
            "部門コード", "部門名", "取引先コード", "取引先名", "ステータス",
            "承認番号", "当月請求額", "インボイス番号", "軽減税率有無", "稟議書有無",
            "読取方式", "読取精度", "代表ファイル", "検証ステータス", "請求書発行日", "ファイルリンク"
        ]
        
        # スタイル定義
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # ヘッダー書き込み
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # データ取得
        # output_summary と結合して情報を補完
        query = """
            SELECT 
                r.dept_code,
                r.dept_name,
                r.vendor_code,
                r.vendor_name,
                r.status,
                r.approval_no,
                r.detected_amount,
                r.detected_invoice_no,
                r.has_reduced_tax,
                r.has_ringi,
                r.ocr_method,
                r.confidence,
                r.file_name,
                r.detected_date
            FROM invoice_ocr_results r
            WHERE r.run_id = ?
            ORDER BY r.dept_code, r.vendor_code, r.approval_no
        """
        
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(query, (run_id,))
            rows = cursor.fetchall()
            
            for row_idx, row in enumerate(rows, 2):
                # データ準備
                confidence = row["confidence"] if row["confidence"] is not None else 0.0
                verification_status = "OK" if confidence >= 0.9 else "要確認"
                
                file_link = f"{base_url}/api/ocr/files/{row['file_name']}"
                
                data = [
                    row["dept_code"],
                    row["dept_name"],
                    row["vendor_code"],
                    row["vendor_name"],
                    row["status"],
                    row["approval_no"],
                    row["detected_amount"],
                    row["detected_invoice_no"],
                    "有" if row["has_reduced_tax"] else "無",
                    "有" if row["has_ringi"] else "無",
                    row["ocr_method"],
                    f"{confidence:.2f}",
                    row["file_name"],
                    verification_status,
                    row["detected_date"],
                    f'=HYPERLINK("{file_link}", "{row["file_name"]}")'
                ]
                
                # 行書き込み
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # File Link Style
                    if col_idx == 16: # File Link
                        cell.font = Font(color="0000FF", underline="single")
                    
                    # Amount Format
                    if col_idx == 7 and isinstance(value, (int, float)):
                        cell.number_format = '"¥"#,##0'
                    
                    # Validation Status Style
                    if col_idx == 14:
                        if value == "OK":
                            cell.font = Font(color="008800", bold=True)
                        else:
                            cell.font = Font(color="FF0000", bold=True)

        # 列幅自動調整 (簡易)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50) # Max 50

        # メモリバッファに出力
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
