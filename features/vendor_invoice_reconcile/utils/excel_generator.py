from datetime import datetime
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import ReconcileResult, RunInfo

class ExcelGenerator:
    def __init__(self):
        # Styles
        self.FONT_HEADER = Font(bold=True, color="FFFFFF")
        self.FILL_HEADER = PatternFill("solid", fgColor="4472C4")
        self.FILL_OK = PatternFill("solid", fgColor="C6EFCE") # Green
        self.FILL_NG = PatternFill("solid", fgColor="FFC7CE") # Red
        self.FILL_WARN = PatternFill("solid", fgColor="FFEB9C") # Yellow
        self.BORDER_THIN = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def generate(self, 
                 results: List[ReconcileResult], 
                 run_info: RunInfo, 
                 output_dir: str = "output/reconcile") -> str:
        
        wb = Workbook()
        
        # 1. Reconciliation Sheet
        ws_rec = wb.active
        ws_rec.title = "Reconciliation"
        self._write_reconciliation_sheet(ws_rec, results)
        
        # 2. Unmapped Sheet
        ws_unmapped = wb.create_sheet("Unmapped")
        self._write_unmapped_sheet(ws_unmapped, results)
        
        # 3. Run Info Sheet
        ws_info = wb.create_sheet("RunInfo")
        self._write_run_info_sheet(ws_info, run_info)
        
        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reconcile_{run_info.base_month}_{timestamp}.xlsx"
        output_path = Path(output_dir) / filename
        
        if not output_path.parent.exists():
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
        wb.save(output_path)
        return str(output_path)

    def _write_reconciliation_sheet(self, ws, results: List[ReconcileResult]):
        headers = [
            "Status", "Base Month", "Vendor Code", "Vendor Name", 
            "Monthly?", # New
            "Dept Code", "Dept Name", 
            "Invoice Amount (Expected)", "Payment Amount (Actual)", "Diff", "Note"
        ]
        
        # Write Header
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.FONT_HEADER
            cell.fill = self.FILL_HEADER
            cell.alignment = Alignment(horizontal="center")
        
        # Write Data
        row_idx = 2
        for res in results:
            if res.status == "UNMAPPED":
                continue # Skip unmapped in main sheet (or keep separate)

            row = [
                res.status, res.base_month, res.vendor_code, res.vendor_name,
                res.is_monthly, # New
                res.dept_code, res.dept_name,
                res.invoice_amount, res.payment_amount, res.diff_amount, res.note
            ]
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER_THIN
                
                # Styling
                if col_idx == 1: # Status
                    if value == "OK": cell.fill = self.FILL_OK
                    elif value == "MISSING": cell.fill = self.FILL_NG
                    elif value == "EXCESS": cell.fill = self.FILL_WARN
                    elif value == "DIFF": cell.fill = self.FILL_NG
                
                # Number Format
                if col_idx in [7, 8, 9]: # Amounts
                    cell.number_format = "#,##0"

            row_idx += 1
            
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx}"
        
        # Column Widths
        ws.column_dimensions["E"].width = 15 # Dept Code
        ws.column_dimensions["F"].width = 30 # Dept Name
        
    def _write_unmapped_sheet(self, ws, results: List[ReconcileResult]):
        headers = [
            "Vendor Code", "Vendor Name", "Raw Dept Name (From Invoice)", 
            "Total Amount", "Source Files", "Suggested Mapping (Fill Here)"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.FONT_HEADER
            cell.fill = self.FILL_HEADER
        
        row_idx = 2
        for res in results:
            if res.status != "UNMAPPED":
                continue
                
            source_files = set(d.source_file for d in res.details)
            
            row = [
                res.vendor_code, res.vendor_name, res.dept_name, # dept_name holds raw name for unmapped
                res.invoice_amount, ", ".join(source_files), ""
            ]
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER_THIN
                if col_idx == 4:
                    cell.number_format = "#,##0"
            
            row_idx += 1

        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["F"].width = 20

    def _write_run_info_sheet(self, ws, run_info: RunInfo):
        info = [
            ("Run Date", run_info.run_date),
            ("Base Month", run_info.base_month),
            ("Target Vendors", ", ".join(run_info.target_vendors)),
            ("Input Files", str(len(run_info.input_files)))
        ]
        
        for i, (k, v) in enumerate(info):
            ws.cell(row=i+1, column=1, value=k).font = Font(bold=True)
            ws.cell(row=i+1, column=2, value=v)
