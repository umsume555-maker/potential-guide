"""
Excel出力モジュール
openpyxlを使用してチェック結果をExcelファイルに出力
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

# Excelで使用できない不正文字定義
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


# スタイル定義
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
DASH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# OUTPUT_SUMMARYのカラム定義
SUMMARY_COLUMNS = [
    ("総合", "overall_result", 8),
    ("OCR判定", "ocr_match_status", 8),
    ("リンク", "ocr_file_link", 10),
    ("当月請求書", "ocr_amount", 10),
    ("支払金額", "payment_amount", 15),
    ("担当", "assignee", 15),
    ("取引日付", "transaction_date", 12),
    ("取引先コード", "vendor_code", 12),
    ("取引先名", "vendor_name", 25),
    ("申請部門コード", "dept_code", 12),
    ("申請部門名", "dept_name", 20),
    ("予定日判定", "payment_date_result", 10),
    ("支払予定日", "payment_date", 12),
    ("予定日（正）", "payment_date_expected", 12),
    ("税区分判定", "tax_result", 10),
    ("税区分", "tax_category", 8),
    ("税区分（正）", "tax_expected", 10),
    ("科目判定", "account_result", 10),
    ("科目", "account_code", 10),
    ("科目名", "account_name", 20),
    ("科目（正）", "account_expected", 10),
    ("科目名（正）", "account_expected_name", 20),
    ("ズレモレ判定", "anomaly_result", 12),
    ("種別", "anomaly_type", 10),
    ("金額(3M前)", "amount_3m_ago", 12),
    ("個数(3M前)", "count_3m_ago", 10),
    ("金額(2M前)", "amount_2m_ago", 12),
    ("個数(2M前)", "count_2m_ago", 10),
    ("金額(1M前)", "amount_1m_ago", 12),
    ("個数(1M前)", "count_1m_ago", 10),
    ("金額(当月)", "amount_current", 12),
    ("個数(当月)", "count_current", 10),
    ("金額(翌月)", "amount_next", 12),
    ("個数(翌月)", "count_next", 10),
    ("支払先相違", "vendor_payee_result", 10),
    ("支払先コード", "payee_code", 12),
    ("伝票番号", "base_invoice_no", 15),
    ("状況区分", "status", 15),
]

# OUTPUT_DETAILのカラム定義
DETAIL_COLUMNS = [
    ("伝票番号", "invoice_no", 18),
    ("ベース伝票", "base_invoice_no", 15),
    ("枝番", "branch_no", 8),
    ("部門コード", "dept_code", 12),
    ("部門名", "dept_name", 20),
    ("取引先コード", "vendor_code", 12),
    ("取引先名", "vendor_name", 25),
    ("支払先コード", "payee_code", 12),
    ("支払先名", "payee_name", 25),
    ("科目コード", "account_code", 10),
    ("科目名", "account_name", 20),
    ("支払金額", "payment_amount", 15),
    ("税区分", "tax_category", 8),
    ("税区分名", "tax_category_name", 20),
    ("支払予定日", "payment_date", 12),
    ("取引日付", "transaction_date", 12),
    ("状況区分", "status", 10),
]


def apply_judgment_style(cell, value: str):
    """判定値に応じたスタイルを適用"""
    if value == "OK":
        cell.fill = OK_FILL
    elif value == "NG":
        cell.fill = NG_FILL
    elif value == "-":
        cell.fill = DASH_FILL


def clean_text(value: Any) -> Any:
    """Excelに使用できない不正文字を除去"""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def write_sheet(
    ws,
    columns: List[tuple],
    data: List[Dict[str, Any]],
    judgment_columns: Optional[List[str]] = None
):
    """
    シートにデータを書き込む
    
    Args:
        ws: ワークシート
        columns: カラム定義 [(表示名, キー, 幅), ...]
        data: データリスト
        judgment_columns: 判定スタイルを適用するカラムのキーリスト
    """
    if judgment_columns is None:
        judgment_columns = []
    
    # ヘッダー行
    for col_idx, (header, key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # データ行
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, (header, key, width) in enumerate(columns, start=1):
            value = row_data.get(key, "")
            
            # 金額はカンマ区切り
            if "amount" in key and isinstance(value, (int, float)):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = "#,##0"
            else:
                # 部門コード補正: 8桁ゼロ埋め (ユーザー指示)
                if key == "dept_code" and value:
                    try:
                        # 一旦数値にしてからフォーマットすることで、既存の "0123" などを安全に処理
                        value = f"{int(str(value)):08d}"
                    except:
                        pass # 変換できない場合(文字列など)はそのまま
                
                cleaned_value = clean_text(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cleaned_value)
                
                # コード系カラムは文字列として書式設定（ゼロ落ち防止）
                if key in ["dept_code", "vendor_code", "payee_code", "account_code", "tax_category"]:
                     cell.number_format = '@'

            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            
            # 判定カラムにスタイル適用
            if key in judgment_columns and isinstance(value, str):
                apply_judgment_style(cell, value)
    
    # フィルター設定
    if data:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data) + 1}"
    
    # ウィンドウ枠固定
    ws.freeze_panes = "A2"


def write_excel(
    output_path: Path,
    base_month: str,
    summary_data: List[Dict[str, Any]],
    detail_data: List[Dict[str, Any]],
    excluded_vendors: List[Dict[str, Any]] = None,
    run_info: Dict[str, Any] = None
) -> Path:
    """
    チェック結果をExcelファイルに出力
    
    Args:
        output_path: 出力先ディレクトリ
        base_month: 基準月 (YYYY-MM)
        summary_data: OUTPUT_SUMMARYデータ
        detail_data: OUTPUT_DETAILデータ
        excluded_vendors: 除外取引先リスト
        run_info: 実行情報
    
    Returns:
        出力ファイルパス
    """
    wb = Workbook()
    
    # OUTPUT_SUMMARY シート
    ws_summary = wb.active
    ws_summary.title = "OUTPUT_SUMMARY"
    
    # ユーザー要望: モレ(毎月あるのに今月ない)はExcelには出力しない
    filtered_summary = [
        row for row in summary_data 
        if row.get("anomaly_type") != "毎月あるのに今月ない"
    ]
    
    judgment_cols = [
        "overall_result", "ocr_match_status", "vendor_payee_result", "tax_result", 
        "account_result", "payment_date_result", "anomaly_result"
    ]
    write_sheet(ws_summary, SUMMARY_COLUMNS, filtered_summary, judgment_cols)
    
    # OUTPUT_DETAIL シート
    ws_detail = wb.create_sheet("OUTPUT_DETAIL")
    write_sheet(ws_detail, DETAIL_COLUMNS, detail_data)
    
    # EXCLUDED_VENDORS シート
    if excluded_vendors:
        ws_excluded = wb.create_sheet("EXCLUDED_VENDORS")
        excluded_columns = [
            ("取引先コード", "vendor_code", 15),
            ("取引先名", "vendor_name", 30),
            ("除外理由", "reason", 30),
        ]
        write_sheet(ws_excluded, excluded_columns, excluded_vendors)
    
    # RUN_INFO シート（実行情報）
    if run_info:
        ws_info = wb.create_sheet("RUN_INFO")
        
        # 基本情報
        info_items = [
            ("実行ID", run_info.get("run_id", "")),
            ("基準月", run_info.get("base_month", "")),
            ("実行開始", run_info.get("started_at", "")),
            ("実行終了", run_info.get("ended_at", "")),
            ("入力行数", run_info.get("input_rows", 0)),
            ("出力行数", run_info.get("output_rows", 0)),
            ("NG件数", run_info.get("ng_count", 0)),
            ("照合保留件数", run_info.get("hold_count", 0)),
            ("対象外件数", run_info.get("dash_count", 0)),
            ("", ""), # 空行
            ("【システム情報】", ""),
            ("正マスター総数", run_info.get("rule_total", 0)),
            ("正マスター最終更新", run_info.get("rule_updated", "")),
            ("参照DBパス", run_info.get("db_path", "")),
            ("", ""),
            ("【データ健全性診断】", ""),
        ]

        # 診断メトリクス（存在する場合）
        # パーセンテージ表示系のキー
        null_rate_keys = [
            ("Null率: dept_code", "null_rate_dept"),
            ("Null率: vendor_code", "null_rate_vendor"),
            ("Null率: payee_code", "null_rate_payee"),
            ("Null率: payment_date", "null_rate_payment_date"),
            ("Null率: tax_category", "null_rate_tax"),
            ("Null率: account_code", "null_rate_account"),
            ("日付一致率 (Planned==Trade)", "date_match_rate"),
        ]
        
        for label, key in null_rate_keys:
            if key in run_info:
                val = run_info[key]
                info_items.append((label, f"{val:.1%}"))
        
        # 分布情報
        if "dist_tax" in run_info:
            info_items.append(("値分布: tax_category", run_info["dist_tax"]))
        if "dist_account" in run_info:
            info_items.append(("値分布: account_code", run_info["dist_account"]))
            
        # エラー系（あれば）
        if "error_message" in run_info and run_info["error_message"]:
            info_items.append(("", ""))
            info_items.append(("【エラー詳細】", run_info["error_message"]))
        
        if "csv_headers" in run_info:
            info_items.append(("CSVヘッダー(先頭80)", run_info["csv_headers"]))

        for row_idx, (label, value) in enumerate(info_items, start=1):
            ws_info.cell(row=row_idx, column=1, value=label)
            # 値が長すぎる場合の対策などは一旦なしで
            ws_info.cell(row=row_idx, column=2, value=value)
            
        ws_info.column_dimensions["A"].width = 30
        ws_info.column_dimensions["B"].width = 60
    
    # ファイル保存
    output_path.mkdir(parents=True, exist_ok=True)
    file_name = f"OUTPUT_{base_month}.xlsx"
    file_path = output_path / file_name
    wb.save(file_path)
    
    return file_path


if __name__ == "__main__":
    # テスト実行
    test_summary = [
        {
            "overall_result": "OK",
            "dept_code": "1101",
            "dept_name": "○○ホーム",
            "vendor_code": "10001234",
            "vendor_name": "株式会社サンプル",
            "base_invoice_no": "PI2511001811",
            "transaction_date": "2025-10-15",
            "payee_code": "10001234",
            "vendor_payee_result": "OK",
            "payment_amount": 52000,
            "payment_date": "2025-11-28",
            "payment_date_result": "OK",
            "payment_date_expected": "2025-11-28",
            "tax_category": "45",
            "tax_result": "OK",
            "tax_expected": "45",
            "account_code": "82870",
            "account_result": "OK",
            "account_expected": "82870",
            "anomaly_result": "OK",
            "anomaly_type": "",
            "is_monthly": "毎月",
            "status": "未承認",
        },
        {
            "overall_result": "NG",
            "dept_code": "1101",
            "dept_name": "○○ホーム",
            "vendor_code": "10004567",
            "vendor_name": "株式会社サンプル設備",
            "base_invoice_no": "PI2511002190",
            "transaction_date": "2025-10-31",
            "payee_code": "10004567",
            "vendor_payee_result": "OK",
            "payment_amount": 210000,
            "payment_date": "2025-12-02",
            "payment_date_result": "NG",
            "payment_date_expected": "2025-11-28",
            "tax_category": "45",
            "tax_result": "OK",
            "tax_expected": "45",
            "account_code": "82820",
            "account_result": "NG",
            "account_expected": "82750",
            "anomaly_result": "NG",
            "anomaly_type": "ズレ",
            "is_monthly": "",
            "status": "未承認",
        },
    ]
    
    test_detail = [
        {
            "invoice_no": "PI2511001811-0001",
            "base_invoice_no": "PI2511001811",
            "branch_no": "0001",
            "dept_code": "1101",
            "dept_name": "○○ホーム",
            "vendor_code": "10001234",
            "vendor_name": "株式会社サンプル",
            "payee_code": "10001234",
            "payee_name": "株式会社サンプル",
            "account_code": "82870",
            "account_name": "支払手数料",
            "payment_amount": 52000,
            "tax_category": "45",
            "tax_category_name": "課税仕入(10%)",
            "payment_date": "2025-11-28",
            "transaction_date": "2025-10-15",
            "status": "未承認",
        }
    ]
    
    from pathlib import Path
    output_dir = Path(__file__).parent.parent / "data"
    result = write_excel(
        output_dir,
        "2025-10",
        test_summary,
        test_detail,
        run_info={
            "run_id": "TEST001",
            "base_month": "2025-10",
            "started_at": "2025-10-15 10:00:00",
            "ended_at": "2025-10-15 10:01:00",
            "input_rows": 100,
            "output_rows": 50,
            "ng_count": 5,
        }
    )
    print(f"出力完了: {result}")
